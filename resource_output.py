#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
AWS 인프라 구성 집계 스크립트 (Python 3.13)
- 계정명 / 지역 / 서비스 / 타입 / 총 갯수 / 동작 상태별 갯수
- 리전 순회 자원 + 글로벌 자원 집계
- 콘솔 요약 출력 + Excel 저장

포함 서비스:
EC2, EBS, S3, RDS, ElastiCache(Redis), DynamoDB, Route53, ELB(ALB/NLB/Classic),
CloudFront, Lambda, EKS, ECS, ECR, API Gateway(v1/v2), SQS, SNS,
OpenSearch(Service/Serverless)

pip install boto3 botocore openpyxl
"""


import sys
from collections import defaultdict, Counter
from datetime import datetime, timezone


import boto3
from botocore.config import Config
from botocore.exceptions import ClientError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from botocore.exceptions import ClientError # 권한 없는 서비스 건너 뛰고 수행용
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime, timezone

import time
from datetime import datetime, timezone

from collections import defaultdict, Counter

from datetime import datetime, timezone

# ============================================================
# 0) 자격증명 하드코딩
# ============================================================
ACCESS_KEY = "엑세스키"
SECRET_KEY = "시크릿키"
SESSION_TOKEN = None  # 필요 없으면 None
ACCOUNT_ALIAS_OR_NAME = "AWS Account명(숫자)"
# ============================================================
# 1) 공통 설정
# ============================================================
BOTO_CONFIG = Config(
    retries={"max_attempts": 10, "mode": "standard"},
    connect_timeout=10,
    read_timeout=60,
    user_agent_extra="infra-inventory/1.2-region-col"
)

def make_session():
    return boto3.Session(
        aws_access_key_id=ACCESS_KEY,
        aws_secret_access_key=SECRET_KEY,
        aws_session_token=SESSION_TOKEN
    )

def list_all_regions(session):
    """상용 리전 리스트"""
    ec2 = session.client("ec2", region_name="us-east-1", config=BOTO_CONFIG)
    resp = ec2.describe_regions(AllRegions=False)
    return sorted([r["RegionName"] for r in resp["Regions"]])

# ============================================================
# 2) 헬퍼
# ============================================================
def safe_get(client_call, **kwargs):
    try:
        return client_call(**kwargs)
    except ClientError as e:
        print(f"[WARN] API error: {e}", file=sys.stderr)
        return {}
    except Exception as e:
        print(f"[WARN] Error: {e}", file=sys.stderr)
        return {}

def dict_counter_add(counter: Counter, key: str, inc: int = 1):
    counter[key] += inc

def summarize_counter(counter: Counter) -> tuple[int, str]:
    total = sum(counter.values())
    parts = [f"{k}:{v}" for k, v in sorted(counter.items())]
    return total, " | ".join(parts) if parts else ""

#아래 4개 함수는 권한이 없어도 스크립트가 중간에 죽지않고 가능한 항목만 집계하도록 방어로직
def _is_access_denied(e: ClientError) -> bool:
    try:
        code = e.response.get("Error", {}).get("Code", "")
    except Exception:
        code = ""
    return code in (
        "UnauthorizedOperation",
        "AccessDenied",
        "AccessDeniedException",
        "UnrecognizedClientException",
        "ExpiredTokenException"
    )

def _is_access_denied_exc(e: Exception) -> bool:
    if not isinstance(e, ClientError):
        return False
    code = e.response.get("Error", {}).get("Code", "")
    return code in {
        "UnauthorizedOperation",
        "AccessDenied",
        "AccessDeniedException",
        "UnrecognizedClientException",
        "ExpiredTokenException",
        "AccessDeniedForDependencyException",
    }

def safe_paginate(paginator, *, on_skip:str="", **kwargs):
    """
    안전한 페이지네이터.
    - 권한/토큰 오류: 경고만 출력하고 '빈 반복'을 반환하여 상위 로직이 계속 진행되도록.
    - 다른 오류: 그대로 raise (디버깅을 위해)
    """
    try:
        for page in paginator.paginate(**kwargs):
            yield page
    except Exception as e:
        if _is_access_denied_exc(e):
            msg = e.response.get("Error", {}).get("Message", str(e))
            print(f"[SKIP] {on_skip or 'paginate'}: AccessDenied → {msg}")
            return
        raise

def safe_call(fn, *, on_skip:str="", **kwargs):
    """
    단건 API 안전 호출.
    - 권한/토큰 오류: 경고 출력 후 빈 dict 반환
    - 다른 오류: 그대로 raise
    """
    try:
        return fn(**kwargs)
    except Exception as e:
        if _is_access_denied_exc(e):
            msg = getattr(e, "response", {}).get("Error", {}).get("Message", str(e))
            print(f"[SKIP] {on_skip or getattr(fn, '__name__', 'call')}: AccessDenied → {msg}")
            return {}
        raise

# for page in paginator.paginate(): -> for page in safe_paginate(paginator): 대체

def _split_status_details(details: str) -> list[tuple[str, int]]:
    """
    'running:3 | stopped:1' → [("running",3), ("stopped",1)]
    그 외(빈 문자열/상태 없음) → []
    """
    if not details:
        return []
    parts = [p.strip() for p in details.split("|")]
    kv = []
    for p in parts:
        if not p:
            continue
        if ":" in p:
            k, v = p.split(":", 1)
            k = k.strip()
            try:
                v = int(v.strip())
            except ValueError:
                # status:ACTIVE 같은 1개 값일 수도 있으니 숫자 아님 → 1로 간주
                v = 1 if p.strip() else 0
            kv.append((k, v))
        else:
            # 'status:ACTIVE' 처럼 콜론이 하나만인 케이스 방어
            # 또는 'Enabled' 단독 같은 값 → 1로 간주
            kv.append((p, 1))
    # 0 값 필터링
    return [(k, v) for k, v in kv if v > 0]

def _label_from_tags(tags: list[dict]) -> str:
    """노드 라벨 결정 우선순위"""
    m = {t.get("Key"): t.get("Value") for t in (tags or [])}
    return (
        m.get("eks:nodegroup-name")
        or m.get("karpenter.sh/provisioner-name")
        or m.get("aws:autoscaling:groupName")
        or "unknown"
    )

def _split_status_details_pairs(details: str) -> list[tuple[str, str]]:
    """
    'running:3 | stopped:1' → [('running','3'), ('stopped','1')]
    'status:ACTIVE'         → [('status:ACTIVE','')]   # 비정수 값은 갯수 공란
    'Enabled'               → [('Enabled','')]         # 키:값 형태가 아니면 상태만
    공백/빈값               → []
    """
    if not details or not isinstance(details, str):
        return []
    parts = [p.strip() for p in details.split("|") if p.strip()]
    pairs: list[tuple[str, str]] = []
    for p in parts:
        if ":" in p:
            k, v = p.split(":", 1)
            k = k.strip()
            v = v.strip()
            # v가 숫자면 갯수로, 아니면 상태 문자열로만 취급
            if v.isdigit():
                pairs.append((k, v))
            else:
                pairs.append((f"{k}:{v}", ""))  # 상태만, 갯수 공란
        else:
            pairs.append((p, ""))  # 상태만, 갯수 공란
    return pairs





# ============================================================
# 3) 수집기 (리전형 / 글로벌형)
#   각 함수는 (계정명, 지역, 서비스, 타입, 총, 상세) 튜플 리스트를 반환
# ============================================================

def collect_ec2(session, regions):
    rows = []
    for region in regions:
        ec2 = session.client("ec2", region_name=region, config=BOTO_CONFIG)
        # 타입별 상태 카운트
        type_state = defaultdict(Counter)
        paginator = ec2.get_paginator("describe_instances")
        for page in safe_paginate(paginator, on_skip=f"EC2[{region}].describe_instances"):
            for r in page.get("Reservations", []):
                for i in r.get("Instances", []):
                    itype = i.get("InstanceType", "unknown")
                    state = i.get("State", {}).get("Name", "unknown")
                    type_state[itype][state] += 1
        for itype, cnt in sorted(type_state.items()):
            total = sum(cnt.values())
            if not total:
                continue
            details = " | ".join(f"{k}:{v}" for k, v in sorted(cnt.items()))
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "EC2", itype, total, details))
            print(f"[EC2][{region}][{itype}] total={total}, {details}")
    return rows


def collect_ebs(session, regions):
    rows = []
    for region in regions:
        ec2 = session.client("ec2", region_name=region, config=BOTO_CONFIG)
        type_state = defaultdict(Counter)
        paginator = ec2.get_paginator("describe_volumes")
        for page in safe_paginate(paginator, on_skip=f"EBS[{region}].describe_volumes"):
            for v in page.get("Volumes", []):
                vtype = v.get("VolumeType", "unknown")   # gp2/gp3/io1/io2/st1/sc1/standard
                state = v.get("State", "unknown")        # in-use/available
                type_state[vtype][state] += 1
        for vtype, cnt in sorted(type_state.items()):
            total = sum(cnt.values())
            if not total:
                continue
            details = " | ".join(f"{k}:{v}" for k, v in sorted(cnt.items()))
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "EBS", vtype, total, details))
            print(f"[EBS][{region}][{vtype}] total={total}, {details}")
    return rows

def collect_s3(session):
    region = "global"
    s3 = session.client("s3", config=BOTO_CONFIG)
    counter = Counter()
    resp = safe_get(s3.list_buckets)
    for b in resp.get("Buckets", []):
        name = b["Name"]
        ver = safe_get(s3.get_bucket_versioning, Bucket=name)
        status = ver.get("Status", "Unversioned")  # Enabled / Suspended / Unversioned
        dict_counter_add(counter, status)
    total, details = summarize_counter(counter)
    rows = []
    if total:
        rows.append((ACCOUNT_ALIAS_OR_NAME, region, "S3", "Buckets", total, details))
        print(f"[S3][{region}] total={total}, {details}")
    return rows

def collect_rds(session, regions):
    rows = []
    for region in regions:
        rds = session.client("rds", region_name=region, config=BOTO_CONFIG)
        cls_state = defaultdict(Counter)  # db.t3.medium -> {available:n, ...}
        paginator = rds.get_paginator("describe_db_instances")
        for page in safe_paginate(paginator, on_skip=f"RDS[{region}].describe_db_instances"):
            for db in page.get("DBInstances", []):
                dbclass = db.get("DBInstanceClass", "unknown")
                status = db.get("DBInstanceStatus", "unknown")
                cls_state[dbclass][status] += 1
        for dbclass, cnt in sorted(cls_state.items()):
            total = sum(cnt.values())
            if not total:
                continue
            details = " | ".join(f"{k}:{v}" for k, v in sorted(cnt.items()))
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "RDS", dbclass, total, details))
            print(f"[RDS][{region}][{dbclass}] total={total}, {details}")
    return rows


def collect_elasticache_redis(session, regions):
    rows = []
    for region in regions:
        ec = session.client("elasticache", region_name=region, config=BOTO_CONFIG)
        type_state = defaultdict(Counter)
        paginator = ec.get_paginator("describe_cache_clusters")
        for page in safe_paginate(paginator, on_skip=f"ElastiCache[{region}].describe_cache_clusters"):
            for cl in page.get("CacheClusters", []):
                if cl.get("Engine", "").lower() != "redis":
                    continue
                ctype = cl.get("CacheNodeType", "unknown")         # cache.t3.micro 등
                status = cl.get("CacheClusterStatus", "unknown")   # available 등
                type_state[ctype][status] += 1
        for ctype, cnt in sorted(type_state.items()):
            total = sum(cnt.values())
            if not total:
                continue
            details = " | ".join(f"{k}:{v}" for k, v in sorted(cnt.items()))
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "ElastiCache", ctype, total, details))
            print(f"[ElastiCache-Redis][{region}][{ctype}] total={total}, {details}")
    return rows


def collect_dynamodb(session, regions):
    rows = []
    for region in regions:
        ddb = session.client("dynamodb", region_name=region, config=BOTO_CONFIG)
        state_counter = Counter()
        billing_counter = Counter()
        paginator = ddb.get_paginator("list_tables")
        for page in safe_paginate(paginator):
            for name in page.get("TableNames", []):
                desc = safe_get(ddb.describe_table, TableName=name)
                td = desc.get("Table", {})
                state = td.get("TableStatus", "unknown")
                dict_counter_add(state_counter, state)
                billing = td.get("BillingModeSummary", {}).get("BillingMode", "PROVISIONED")
                dict_counter_add(billing_counter, billing)
        total_s, details_s = summarize_counter(state_counter)
        total_b, details_b = summarize_counter(billing_counter)
        if total_s:
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "DynamoDB", "Tables", total_s, details_s))
            print(f"[DynamoDB][{region}] tables={total_s}, states: {details_s}")
        if total_b:
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "DynamoDB", "BillingMode", total_b, details_b))
            print(f"[DynamoDB][{region}] billing: {details_b}")
    return rows

def collect_route53(session):
    region = "global"
    r53 = session.client("route53", config=BOTO_CONFIG)
    counter = Counter()
    paginator = r53.get_paginator("list_hosted_zones")
    for page in safe_paginate(paginator):
        for hz in page.get("HostedZones", []):
            private = hz.get("Config", {}).get("PrivateZone", False)
            dict_counter_add(counter, "Private" if private else "Public")
    total, details = summarize_counter(counter)
    rows = []
    if total:
        rows.append((ACCOUNT_ALIAS_OR_NAME, region, "Route53", "HostedZones", total, details))
        print(f"[Route53][{region}] zones={total}, {details}")
    return rows

def collect_elb(session, regions):
    rows = []
    for region in regions:
        # ---- ELBv2 (ALB/NLB/GWLB) ----
        elbv2 = session.client("elbv2", region_name=region, config=BOTO_CONFIG)
        v2_counter = defaultdict(Counter)

        try:
            paginator = elbv2.get_paginator("describe_load_balancers")
            for page in safe_paginate(paginator, on_skip=f"ELBv2[{region}].describe_load_balancers"):
                for lb in page.get("LoadBalancers", []):
                    lb_type = lb.get("Type", "unknown")   # application / network / gateway
                    state = lb.get("State", {}).get("Code", "unknown")
                    v2_counter[lb_type][state] += 1
        except Exception as e:
            # 비권한 오류 외의 예외는 한 리전만 스킵하고 다음 리전 계속
            print(f"[WARN] ELBv2[{region}] unexpected error: {e}")

        for t, c in v2_counter.items():
            total = sum(c.values())
            if not total:
                continue
            type_label = "ALB" if t == "application" else ("NLB" if t == "network" else ("GWLB" if t == "gateway" else t))
            details = " | ".join(f"{k}:{v}" for k, v in sorted(c.items()))
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "ELB", type_label, total, details))
            print(f"[ELBv2][{region}][{type_label}] total={total}, {details}")

        # ---- Classic ELB ----
        elb = session.client("elb", region_name=region, config=BOTO_CONFIG)
        c_counter = Counter()
        try:
            paginator2 = elb.get_paginator("describe_load_balancers")
            for page in safe_paginate(paginator2, on_skip=f"ELB-Classic[{region}].describe_load_balancers"):
                # Classic은 상태코드가 별도 노출되지 않으므로 presence만 집계
                for _ in page.get("LoadBalancerDescriptions", []):
                    c_counter["present"] += 1
        except Exception as e:
            print(f"[WARN] ELB-Classic[{region}] unexpected error: {e}")

        if sum(c_counter.values()) > 0:
            total = sum(c_counter.values())
            details = " | ".join(f"{k}:{v}" for k, v in sorted(c_counter.items()))
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "ELB", "Classic", total, details))
            print(f"[ELB-Classic][{region}] total={total}, {details}")

    return rows


def collect_cloudfront(session):
    region = "global"
    cf = session.client("cloudfront", config=BOTO_CONFIG)
    counter = Counter()
    paginator = cf.get_paginator("list_distributions")
    for page in safe_paginate(paginator):
        dist_list = page.get("DistributionList", {})
        for item in dist_list.get("Items", []) or []:
            enabled = item.get("Enabled", False)
            dict_counter_add(counter, "Enabled" if enabled else "Disabled")
    total, details = summarize_counter(counter)
    rows = []
    if total:
        rows.append((ACCOUNT_ALIAS_OR_NAME, region, "CloudFront", "Distributions", total, details))
        print(f"[CloudFront][{region}] total={total}, {details}")
    return rows

def collect_lambda(session, regions):
    rows = []
    for region in regions:
        lm = session.client("lambda", region_name=region, config=BOTO_CONFIG)
        runtime_counter = Counter()
        state_counter = Counter()
        paginator = lm.get_paginator("list_functions")
        for page in safe_paginate(paginator):
            for f in page.get("Functions", []):
                runtime = f.get("Runtime", "unknown")
                dict_counter_add(runtime_counter, runtime)
                state = f.get("State", "Unknown")
                dict_counter_add(state_counter, state)
        total_r, details_r = summarize_counter(runtime_counter)
        total_s, details_s = summarize_counter(state_counter)
        if total_r:
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "Lambda", "Runtime", total_r, details_r))
            print(f"[Lambda][{region}] runtimes={total_r}, {details_r}")
        if total_s:
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "Lambda", "State", total_s, details_s))
            print(f"[Lambda][{region}] states={details_s}")
    return rows

def collect_eks(session, regions):
    """
    EKS 클러스터별 행 생성
    - 타입 : "<clusterName> | v<version>"
    - 총 갯수 : 1 (클러스터 단위)
    - 상세 : status:<STATUS>
    """
    rows = []
    for region in regions:
        eks = session.client("eks", region_name=region, config=BOTO_CONFIG)
        paginator = eks.get_paginator("list_clusters")
        for page in safe_paginate(paginator, on_skip=f"EKS[{region}].list_clusters"):
            for name in page.get("clusters", []):
                desc = safe_call(eks.describe_cluster, on_skip=f"EKS[{region}].describe_cluster", name=name)
                cl = desc.get("cluster") or {}
                ver = cl.get("version", "unknown")
                status = cl.get("status", "unknown")
                type_label = f"{name} | v{ver}"
                rows.append((
                    ACCOUNT_ALIAS_OR_NAME,
                    region,
                    "EKS",
                    type_label,
                    1,                      # 클러스터 1개
                    f"status:{status}"
                ))
                print(f"[EKS][{region}][{name}] version=v{ver}, status={status}")
    return rows

def collect_eks_tree(session, regions):
    """
    EKS 트리 출력:
    - 1) EKS(클러스터) 행 : 타입 "<cluster> | v<version>", 총=1, 상세 "status:ACTIVE"
    - 2) 바로 아래 노드(서비스명 공백) 행들 : 타입 "<nodeLabel> | <instanceType>", 총=노드수, 상세 "running:n | ..."
    """
    rows = []
    for region in regions:
        eks = session.client("eks", region_name=region, config=BOTO_CONFIG)
        ec2 = session.client("ec2", region_name=region, config=BOTO_CONFIG)

        paginator = eks.get_paginator("list_clusters")
        for page in safe_paginate(paginator, on_skip=f"EKS[{region}].list_clusters"):
            for cluster in page.get("clusters", []):
                # --- 클러스터 행
                desc = safe_call(eks.describe_cluster, on_skip=f"EKS[{region}].describe_cluster", name=cluster)
                cl = desc.get("cluster") or {}
                ver = cl.get("version", "unknown")
                status = cl.get("status", "unknown")
                type_label = f"{cluster} | v{ver}"
                rows.append((ACCOUNT_ALIAS_OR_NAME, region, "EKS", type_label, 1, f"status:{status}"))
                print(f"[EKS][{region}][{cluster}] version=v{ver}, status={status}")

                # --- 노드 집계(EC2)
                filters = [
                    {"Name": f"tag:kubernetes.io/cluster/{cluster}", "Values": ["owned", "shared"]},
                ]
                type_state_by_node_label: dict[tuple[str, str], Counter] = defaultdict(Counter)
                inst_p = ec2.get_paginator("describe_instances")
                for ipage in safe_paginate(inst_p,
                                           on_skip=f"EKS-Nodes[{region}].describe_instances:{cluster}",
                                           Filters=filters):
                    for res in ipage.get("Reservations", []):
                        for inst in res.get("Instances", []):
                            itype = inst.get("InstanceType", "unknown")
                            state = (inst.get("State") or {}).get("Name", "unknown")
                            node_label = _label_from_tags(inst.get("Tags", []))
                            type_state_by_node_label[(node_label, itype)][state] += 1

                # 노드 행 추가 (서비스명 공백으로 트리 시각화)
                for (node_label, itype), cnt in sorted(type_state_by_node_label.items()):
                    total = sum(cnt.values())
                    if not total:
                        continue
                    details = " | ".join(f"{k}:{v}" for k, v in sorted(cnt.items()))
                    typelabel = f"{node_label} | {itype}"
                    rows.append(("", region, "", typelabel, total, details))
                    print(f"[EKS-Nodes][{region}][{cluster}][{node_label}][{itype}] total={total}, {details}")

    return rows


from collections import defaultdict, Counter

def collect_eks_nodes(session, regions):
    """
    EKS 노드 집계 (EC2 인스턴스 태그 기반)
    - 태그: kubernetes.io/cluster/<CLUSTER_NAME> = owned | shared
    - 타입 : "<clusterName> | <instanceType>" (예: "prod-cluster | c5.xlarge")
    - 총 갯수 : 해당 타입 노드 수
    - 상세 : 상태별 카운트 (running, stopped 등)
    """
    rows = []
    for region in regions:
        eks = session.client("eks", region_name=region, config=BOTO_CONFIG)
        ec2 = session.client("ec2", region_name=region, config=BOTO_CONFIG)

        # cluster -> instanceType -> stateCounter
        cluster_type_state: dict[str, dict[str, Counter]] = defaultdict(lambda: defaultdict(Counter))

        # 1) 클러스터 목록
        paginator = eks.get_paginator("list_clusters")
        for page in safe_paginate(paginator, on_skip=f"EKS[{region}].list_clusters"):
            for cluster in page.get("clusters", []):
                # 2) 클러스터 노드(EC2) 조회
                filters = [
                    {"Name": f"tag:kubernetes.io/cluster/{cluster}", "Values": ["owned", "shared"]},
                    # 필요 시 상태 제한 (예: running만) → 아래 주석 해제
                    # {"Name": "instance-state-name", "Values": ["pending", "running", "stopping", "stopped"]}
                ]
                inst_p = ec2.get_paginator("describe_instances")
                for ipage in safe_paginate(inst_p,
                                           on_skip=f"EKS-Nodes[{region}].describe_instances:{cluster}",
                                           Filters=filters):
                    for res in ipage.get("Reservations", []):
                        for inst in res.get("Instances", []):
                            itype = inst.get("InstanceType", "unknown")
                            state = (inst.get("State") or {}).get("Name", "unknown")
                            cluster_type_state[cluster][itype][state] += 1

        # 3) 행 생성
        for cluster, type_map in sorted(cluster_type_state.items()):
            for itype, cnt in sorted(type_map.items()):
                total = sum(cnt.values())
                if not total:
                    continue
                details = " | ".join(f"{k}:{v}" for k, v in sorted(cnt.items()))
                type_label = f"{cluster} | {itype}"
                rows.append((
                    ACCOUNT_ALIAS_OR_NAME,
                    region,
                    "EKS-Nodes",
                    type_label,
                    total,
                    details
                ))
                print(f"[EKS-Nodes][{region}][{cluster}][{itype}] total={total}, {details}")
    return rows



def collect_ecs(session, regions):
    rows = []
    for region in regions:
        ecs = session.client("ecs", region_name=region, config=BOTO_CONFIG)

        # Clusters(기존 유지)
        cluster_state = Counter()
        clusters = []
        paginator = ecs.get_paginator("list_clusters")
        for page in safe_paginate(paginator, on_skip=f"ECS[{region}].list_clusters"):
            clusters += page.get("clusterArns", [])

        if clusters:
            for i in range(0, len(clusters), 100):
                part = clusters[i:i+100]
                resp = safe_call(ecs.describe_clusters, on_skip=f"ECS[{region}].describe_clusters", clusters=part, include=['STATISTICS'])
                for c in (resp.get("clusters") or []):
                    cluster_state[c.get("status","UNKNOWN")] += 1
        if sum(cluster_state.values()):
            total = sum(cluster_state.values())
            details = " | ".join(f"{k}:{v}" for k, v in sorted(cluster_state.items()))
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "ECS", "Clusters", total, details))
            print(f"[ECS][{region}][Clusters] total={total}, {details}")

        # Services : LaunchType 별
        lt_state = defaultdict(Counter)
        for arn in clusters:
            svc_arns = []
            paginator = ecs.get_paginator("list_services")
            for page in safe_paginate(paginator, on_skip=f"ECS[{region}].list_services", cluster=arn):
                svc_arns += page.get("serviceArns", [])
            for i in range(0, len(svc_arns), 10):
                part = svc_arns[i:i+10]
                resp = safe_call(ecs.describe_services, on_skip=f"ECS[{region}].describe_services", cluster=arn, services=part)
                for s in (resp.get("services") or []):
                    lt = (s.get("launchType") or "UNKNOWN").upper()  # EC2/FARGATE/EXTERNAL
                    status = s.get("status", "UNKNOWN")
                    lt_state[lt][status] += 1

        for lt, cnt in sorted(lt_state.items()):
            total = sum(cnt.values())
            if not total:
                continue
            details = " | ".join(f"{k}:{v}" for k, v in sorted(cnt.items()))
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "ECS", lt, total, details))
            print(f"[ECS][{region}][{lt}] total={total}, {details}")
    return rows


def collect_ecr(session, regions):
    rows = []
    for region in regions:
        ecr = session.client("ecr", region_name=region, config=BOTO_CONFIG)
        # 리포지토리
        repos = []
        paginator = ecr.get_paginator("describe_repositories")
        for page in safe_paginate(paginator):
            repos += page.get("repositories", [])
        repo_total = len(repos)
        if repo_total:
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "ECR", "Repositories", repo_total, ""))
            print(f"[ECR][{region}] repositories={repo_total}")

        # 이미지 (TAGGED / UNTAGGED)
        img_counter = Counter()
        for r in repos:
            name = r.get("repositoryName")
            # TAGGED
            paginator_t = ecr.get_paginator("list_images")
            for page in paginator_t.paginate(repositoryName=name, filter={"tagStatus": "TAGGED"}):
                img_counter["TAGGED"] += len(page.get("imageIds", []))
            # UNTAGGED
            paginator_u = ecr.get_paginator("list_images")
            for page in paginator_u.paginate(repositoryName=name, filter={"tagStatus": "UNTAGGED"}):
                img_counter["UNTAGGED"] += len(page.get("imageIds", []))
        if sum(img_counter.values()):
            total, details = summarize_counter(img_counter)
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "ECR", "Images", total, details))
            print(f"[ECR][{region}] images={total}, {details}")
    return rows

def collect_apigw(session, regions):
    rows = []
    for region in regions:
        # v1 (REST)
        apigw = session.client("apigateway", region_name=region, config=BOTO_CONFIG)
        rest_total = 0
        paginator = apigw.get_paginator("get_rest_apis")
        for page in safe_paginate(paginator):
            rest_total += len(page.get("items", []))
        if rest_total:
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "API Gateway", "REST", rest_total, "REST:{}".format(rest_total)))
            print(f"[APIGWv1][{region}] REST total={rest_total}")

        # v2 (HTTP/WEBSOCKET)
        apigw2 = session.client("apigatewayv2", region_name=region, config=BOTO_CONFIG)
        proto_counter = Counter()
        paginator2 = apigw2.get_paginator("list_apis")
        for page in paginator2.paginate():
            for api in page.get("Items", []):
                proto = api.get("ProtocolType", "UNKNOWN")  # HTTP/WEBSOCKET
                dict_counter_add(proto_counter, proto)
        if sum(proto_counter.values()):
            total, details = summarize_counter(proto_counter)
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "API Gateway", "v2", total, details))
            print(f"[APIGWv2][{region}] {details}")
    return rows

def collect_sqs(session, regions):
    rows = []
    for region in regions:
        sqs = session.client("sqs", region_name=region, config=BOTO_CONFIG)
        urls = []
        resp = safe_get(sqs.list_queues)
        if resp and resp.get("QueueUrls"):
            urls = resp["QueueUrls"]
        counter = Counter()
        for u in urls:
            if u.endswith(".fifo"):
                counter["FIFO"] += 1
            else:
                counter["Standard"] += 1
        if sum(counter.values()):
            total, details = summarize_counter(counter)
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "SQS", "Queues", total, details))
            print(f"[SQS][{region}] total={total}, {details}")
    return rows

def collect_sns(session, regions):
    rows = []
    for region in regions:
        sns = session.client("sns", region_name=region, config=BOTO_CONFIG)
        total = 0
        paginator = sns.get_paginator("list_topics")
        for page in safe_paginate(paginator):
            total += len(page.get("Topics", []))
        if total:
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "SNS", "Topics", total, f"Topics:{total}"))
            print(f"[SNS][{region}] topics={total}")
    return rows

def collect_opensearch(session, regions):
    rows = []
    for region in regions:
        # OpenSearch Service (도메인)
        oss = session.client("opensearch", region_name=region, config=BOTO_CONFIG)
        names_resp = safe_get(oss.list_domain_names)
        domain_names = [d.get("DomainName") for d in names_resp.get("DomainNames", [])]
        state_counter = Counter()
        version_counter = Counter()
        if domain_names:
            for name in domain_names:
                d = safe_get(oss.describe_domain, DomainName=name)
                st = (d.get("DomainStatus") or {})
                processing = st.get("Processing", False)
                endpoint = st.get("Endpoint")
                state = "Processing" if processing else ("Active" if endpoint else "Created")
                dict_counter_add(state_counter, state)
                eng = st.get("EngineVersion", "unknown")
                dict_counter_add(version_counter, eng)
        if sum(state_counter.values()):
            total, details = summarize_counter(state_counter)
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "OpenSearch", "Domains", total, details))
            print(f"[OpenSearch@Service][{region}] domains={total}, {details}")
        if sum(version_counter.values()):
            total, details = summarize_counter(version_counter)
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "OpenSearch", "EngineVersion", total, details))
            print(f"[OpenSearch@Service][{region}] versions={details}")

        # OpenSearch Serverless (컬렉션)
        ossv = session.client("opensearchserverless", region_name=region, config=BOTO_CONFIG)
        coll_counter = Counter()
        paginator = ossv.get_paginator("list_collections")
        for page in safe_paginate(paginator):
            for c in page.get("collectionSummaries", []):
                status = c.get("status", "UNKNOWN")
                dict_counter_add(coll_counter, status)
        if sum(coll_counter.values()):
            total, details = summarize_counter(coll_counter)
            rows.append((ACCOUNT_ALIAS_OR_NAME, region, "OpenSearch-Serverless", "Collections", total, details))
            print(f"[OpenSearch@Serverless][{region}] collections={total}, {details}")
    return rows

# ============================================================
# 4) 엑셀 저장
# ============================================================
def save_to_excel(rows, outfile):
    """
    rows: (계정명, 지역, 서비스명, 타입, 총 갯수, 상세문자열) 튜플 리스트
    - 상세문자열을 '상태/상태별 갯수' 2열로 분해
    - 상태가 여러개면 행으로 분리, 2행째부터는 앞 열(계정명/지역/서비스명/타입/총 갯수)은 공란
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = ["계정명", "지역", "서비스명", "타입", "총 갯수", "상태", "상태별 갯수", "생성시각(UTC)"]
    ws.append(headers)

    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    for acc, region, svc, typ, total, details in rows:
        pairs = _split_status_details_pairs(details)
        if not pairs:
            # 상태가 없으면 상태/갯수 공란으로 단일 행
            ws.append([acc, region, svc, typ, total, "", "", ts])
            continue

        # 첫 상태 행: 모든 컬럼 채움
        s0, c0 = pairs[0]
        ws.append([acc, region, svc, typ, total, s0, c0, ts])

        # 이후 상태 행: 앞 열 공란, 상태/갯수만 채움
        for s, c in pairs[1:]:
            ws.append(["", "", "", "", "", s, c, ""])

    # 보기 좋게 너비/정렬
    col_widths = [18, 16, 22, 30, 10, 20, 14, 20]
    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in r:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(outfile)
    print(f"[OK] Excel 저장 완료 : {outfile}")

def _run_collect(collector, session, regions_or_none, sink_rows, label):
    start = time.perf_counter()
    before = len(sink_rows)
    try:
        if regions_or_none is None:
            sink_rows += collector(session)              # 글로벌형
        else:
            sink_rows += collector(session, regions_or_none)  # 리전형
    except Exception as e:
        print(f"[WARN] collector {label} failed (continue): {e}")
    finally:
        elapsed = time.perf_counter() - start
        added = len(sink_rows) - before
        print(f"[TIME] {label} elapsed={elapsed:.3f}s, rows+={added}")


# ============================================================
# 5) 메인
# ============================================================
def main():
    t_total = time.perf_counter()
    started_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S %Z")
    print(f"[INFO] 시작(UTC) : {started_at}")

    print("[INFO] 세션 생성…")
    session = make_session()

    print("[INFO] 리전 조회…")
    regions = list_all_regions(session)

    all_rows = []

    # 리전형
    _run_collect(collect_ec2, session, regions, all_rows, "EC2")
    _run_collect(collect_ebs, session, regions, all_rows, "EBS")
    _run_collect(collect_rds, session, regions, all_rows, "RDS")
    _run_collect(collect_elasticache_redis, session, regions, all_rows, "ElastiCache-Redis")
    _run_collect(collect_dynamodb, session, regions, all_rows, "DynamoDB")
    _run_collect(collect_elb, session, regions, all_rows, "ELB")
    _run_collect(collect_lambda, session, regions, all_rows, "Lambda")
    #_run_collect(collect_eks, session, regions, all_rows, "EKS")
    #_run_collect(collect_eks_nodes, session, regions, all_rows, "EKS-Nodes")
    _run_collect(collect_eks_tree, session, regions, all_rows, "EKS-Tree")
    _run_collect(collect_ecs, session, regions, all_rows, "ECS")
    _run_collect(collect_ecr, session, regions, all_rows, "ECR")
    _run_collect(collect_apigw, session, regions, all_rows, "APIGW")
    _run_collect(collect_sqs, session, regions, all_rows, "SQS")
    _run_collect(collect_sns, session, regions, all_rows, "SNS")
    _run_collect(collect_opensearch, session, regions, all_rows, "OpenSearch")

    # 글로벌형
    _run_collect(collect_s3, session, None, all_rows, "S3")
    _run_collect(collect_route53, session, None, all_rows, "Route53")
    _run_collect(collect_cloudfront, session, None, all_rows, "CloudFront")

    if not all_rows:
        print("[WARN] 수집된 리소스가 없습니다. (권한/리전/SCP 확인)")
        return

    out_name = f"aws_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    save_to_excel(all_rows, out_name)

    ended_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S %Z")
    total_elapsed = time.perf_counter() - t_total
    print(f"[INFO] 종료(UTC) : {ended_at}")
    print(f"[TIME] TOTAL elapsed={total_elapsed:.3f}s, rows={len(all_rows)}")

if __name__ == "__main__":
    main()
